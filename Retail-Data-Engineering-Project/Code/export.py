
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "Output", "Retail_Analytics_Output.xlsx")


HEADER_FILL = PatternFill("solid", start_color="1F4E79")   # Dark blue
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")   # Light blue
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
KPI_FILL    = PatternFill("solid", start_color="2E86C1")
ACCENT_FILL = PatternFill("solid", start_color="E8F4FD")

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(name="Arial", bold=True, color="1F4E79", size=14)
NORMAL_FONT  = Font(name="Arial", size=10)
KPI_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=12)
KPI_VAL_FONT = Font(name="Arial", bold=True, color="1F4E79", size=16)

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_header_row(ws, headers: list, start_row: int = 1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_data_rows(ws, df: pd.DataFrame, start_row: int = 2):
    for row_idx, row in enumerate(df.itertuples(index=False), start_row):
        fill = ALT_FILL if (row_idx - start_row) % 2 == 0 else WHITE_FILL
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")


def _auto_column_width(ws, min_width=12, max_width=40):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)



def write_cleaned_data(ws, df: pd.DataFrame):
    ws.title = "Cleaned_Data"
    title_cell = ws["A1"]
    title_cell.value = "Retail Transactions — Cleaned & Curated Dataset"
    title_cell.font = TITLE_FONT
    ws.row_dimensions[1].height = 25

    _write_header_row(ws, list(df.columns), start_row=2)
    _write_data_rows(ws, df, start_row=3)
    _auto_column_width(ws)
    ws.freeze_panes = "A3"


def write_kpi_summary(ws, df: pd.DataFrame):
    ws.title = "KPI_Summary"
    ws["A1"] = "Executive KPI Summary"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    headers = list(df.columns)
    _write_header_row(ws, headers, start_row=2)

    for row_idx, row in enumerate(df.itertuples(index=False), 3):
        ws.cell(row=row_idx, column=1, value=row[0]).font = Font(name="Arial", bold=True, size=11, color="1F4E79")
        ws.cell(row=row_idx, column=1).fill = ACCENT_FILL
        val_cell = ws.cell(row=row_idx, column=2, value=row[1])
        val_cell.font = KPI_VAL_FONT
        val_cell.fill = WHITE_FILL
        val_cell.alignment = Alignment(horizontal="right")
        for c in [1, 2]:
            ws.cell(row=row_idx, column=c).border = THIN_BORDER

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


def write_generic_table(ws, df: pd.DataFrame, title: str):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 25

    _write_header_row(ws, list(df.columns), start_row=2)
    _write_data_rows(ws, df, start_row=3)
    _auto_column_width(ws)
    ws.freeze_panes = "A3"



def export_to_excel(
    cleaned_df: pd.DataFrame,
    kpis: dict,
    forecast_table: pd.DataFrame,
):
    """
    Write all 6 sheets to Retail_Analytics_Output.xlsx.
    """
    wb = Workbook()

    
    ws1 = wb.active
    write_cleaned_data(ws1, cleaned_df)

    
    ws2 = wb.create_sheet("KPI_Summary")
    write_kpi_summary(ws2, kpis["kpi_summary"])

    
    ws3 = wb.create_sheet("Revenue_By_City")
    write_generic_table(ws3, kpis["revenue_by_city"], "Revenue by City")

  
    ws4 = wb.create_sheet("Revenue_By_Category")
    write_generic_table(ws4, kpis["revenue_by_category"], "Revenue by Category")

   
    ws5 = wb.create_sheet("Top_Products")
    write_generic_table(ws5, kpis["top_products"], "Top Products by Revenue & Quantity")

    ws6 = wb.create_sheet("Sales_Forecast")
    write_generic_table(ws6, forecast_table, "Monthly Sales Forecast (Historical + Predicted)")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    logger.info(f"✅ Excel output saved: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == "__main__":
    print("Run main.py to generate the output file.")
